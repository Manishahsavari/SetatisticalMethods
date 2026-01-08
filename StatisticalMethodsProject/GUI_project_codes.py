import sys
import unicodedata
from collections import defaultdict
import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from scipy import stats
from sklearn.linear_model import LinearRegression
from sklearn.model_selection import train_test_split
import matplotlib.pyplot as plt
import seaborn as sns
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import statsmodels.api as sm

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass


class ExcelReader:
    def __init__(self, path, max_col_limit=22):
        self.path = path
        self.max_col_limit = max_col_limit

    @staticmethod
    def _normalize_cols(cols):
        return [unicodedata.normalize("NFKC", str(c)).strip() for c in cols]

    def _safe_usecols_for_sheet(self, sheet_name):
        try:
            wb = load_workbook(self.path, read_only=True)
            ws = wb[str(sheet_name)]
            max_col = ws.max_column
            last_col_letter = get_column_letter(min(max_col, self.max_col_limit))
            return f"A:{last_col_letter}"
        except Exception:
            return None

    def read(self, sheet_name, usecols=None):
        if usecols is None:
            usecols = self._safe_usecols_for_sheet(sheet_name)
        try:
            if usecols:
                df = pd.read_excel(self.path, sheet_name=str(sheet_name), usecols=usecols, engine="openpyxl")
            else:
                df = pd.read_excel(self.path, sheet_name=str(sheet_name), engine="openpyxl")
            df.columns = self._normalize_cols(df.columns)
            return df
        except Exception:
            try:
                df = pd.read_excel(self.path, sheet_name=str(sheet_name), engine="openpyxl")
                df.columns = self._normalize_cols(df.columns)
                return df
            except Exception as e2:
                print(f"Failed to read sheet {sheet_name}: {e2}")
                return None


class DataAggregator:
    def __init__(self, reader, days=None):
        self.reader = reader
        self.days = days or ["Saturday", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
        self.acc = defaultdict(lambda: defaultdict(list))

    def aggregate(self, sheet_range):
        for sheet in sheet_range:
            df = self.reader.read(sheet_name=str(sheet), usecols="A:V")
            if df is None:
                continue
            if "Stuff" not in df.columns:
                continue
            for _, row in df.iterrows():
                stuff = str(row["Stuff"]).strip()
                if not stuff or stuff.lower() == "nan":
                    continue
                _ = self.acc[stuff]
                row_list = list(row)
                for i, day in enumerate(self.days):
                    start = 1 + i * 3
                    block = row_list[start:start + 3]
                    if all(pd.isna(x) for x in block):
                        continue
                    self.acc[stuff][day].append(block)

    def to_per_stuff_dfs(self):
        per_stuff = {}
        for stuff, daydict in self.acc.items():
            frames = []
            for day in self.days:
                blocks = daydict.get(day, [])
                if not blocks:
                    continue
                df_day = pd.DataFrame(blocks, columns=["Production Plan", "Actual Production", "Success Rate"])
                df_day["Day"] = day
                df_day = df_day[["Day", "Production Plan", "Actual Production", "Success Rate"]]
                frames.append(df_day)
            if frames:
                df_all = pd.concat(frames, ignore_index=True)
                if "Success Rate" in df_all.columns:
                    df_all = df_all.dropna(subset=["Success Rate"]).reset_index(drop=True)
                per_stuff[stuff] = df_all
            else:
                per_stuff[stuff] = pd.DataFrame(columns=["Day", "Production Plan", "Actual Production", "Success Rate"])
        return per_stuff


class Analyzer:
    def __init__(self, df):
        self.df = df.copy()

    def clean_numeric(self):
        for col in ["Production Plan", "Actual Production", "Success Rate"]:
            if col in self.df.columns:
                self.df[col] = pd.to_numeric(self.df[col], errors="coerce")

    def t_test_plan_vs_actual(self, drop_zeros=True):
        df = self.df
        if drop_zeros:
            df = df[df["Actual Production"] != 0]
        a = df["Production Plan"].dropna()
        b = df["Actual Production"].dropna()
        if len(a) < 2 or len(b) < 2:
            return None, None
        t_stat, p_value = stats.ttest_ind(a, b, nan_policy="omit")
        return t_stat, p_value

    def correlation_matrix(self, method="pearson"):
        return self.df.select_dtypes("number").corr(method=method)

    def linear_regression(self, feature="Production Plan", target="Actual Production", test_size=0.2, random_state=42):
        X = self.df[[feature]].dropna()
        y = self.df.loc[X.index, target].dropna()
        X = X.loc[y.index]
        if len(X) < 2:
            return None
        X_train, X_test, y_train, y_test = train_test_split(X.values, y.values, test_size=test_size, random_state=random_state)
        model = LinearRegression().fit(X_train, y_train)
        y_pred = model.predict(X_test)
        return {
            "model": model,
            "X_test": X_test,
            "y_test": y_test,
            "y_pred": y_pred,
            "r2": model.score(X_test, y_test),
            "coef": model.coef_,
            "intercept": model.intercept_,
        }

    def refill_success_rate(self):
        df = self.df
        if "Success Rate" in df.columns and "Production Plan" in df.columns and "Actual Production" in df.columns:
            mask = df["Success Rate"].isna() & df["Production Plan"].notna() & (df["Production Plan"] != 0)
            df.loc[mask, "Success Rate"] = (df.loc[mask, "Actual Production"] / df.loc[mask, "Production Plan"]) * 100
            self.df = df

    def refill_actual_from_success(self, use_success_rate=True, min_samples=5):
        df = self.df.copy()
        required = {"Actual Production", "Production Plan"}
        if not required.issubset(df.columns):
            return 0

        for col in ["Production Plan", "Actual Production", "Success Rate"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")

        target_mask = (
            df["Production Plan"].notna()
            & (df["Production Plan"] != 0)
            & (df["Actual Production"] == 0)
        )

        features = ["Production Plan"]
        if use_success_rate and "Success Rate" in df.columns:
            features.append("Success Rate")

        train_mask = df["Actual Production"].notna() & (df["Actual Production"] != 0) & df[features].notna().all(axis=1)
        n_train = int(train_mask.sum())
        filled = 0

        if n_train >= min_samples:
            X_train = df.loc[train_mask, features].values
            y_train = df.loc[train_mask, "Actual Production"].values
            try:
                model = LinearRegression().fit(X_train, y_train)
                predict_mask = target_mask & df[features].notna().all(axis=1)
                if predict_mask.any():
                    X_pred = df.loc[predict_mask, features].values
                    y_pred = model.predict(X_pred)
                    y_pred = np.where(np.isnan(y_pred), np.nan, y_pred)
                    df.loc[predict_mask, "Actual Production"] = y_pred
                    filled = int(predict_mask.sum())
            except Exception:
                filled = 0
        else:
            if "Success Rate" in df.columns:
                fallback_mask = target_mask & df["Success Rate"].notna()
                if fallback_mask.any():
                    df.loc[fallback_mask, "Actual Production"] = (
                        df.loc[fallback_mask, "Production Plan"] * (df.loc[fallback_mask, "Success Rate"] / 100.0)
                    )
                    filled = int(fallback_mask.sum())

        self.df = df
        return filled


class Visualizer:
    @staticmethod
    def scatter_with_regression(X, y, model=None, title=None, xlabel=None, ylabel=None, ax=None):
        if ax is None:
            fig, ax = plt.subplots(figsize=(8, 6))
        ax.scatter(X, y, color="blue", label="Data points")
        if model is not None:
            xs = np.linspace(np.nanmin(X), np.nanmax(X), 100).reshape(-1, 1)
            ax.plot(xs, model.predict(xs), color="red", linewidth=2, label="Regression line")
        ax.set_xlabel(xlabel or "X")
        ax.set_ylabel(ylabel or "Y")
        if title:
            ax.set_title(title)
        ax.legend()
        return ax

    @staticmethod
    def hist_success_rate(df, ax=None):
        if ax is None:
            fig, ax = plt.subplots(figsize=(8, 6))
        sns.histplot(df["Success Rate"].dropna(), bins=20, kde=True, ax=ax)
        ax.set_title("Distribution of Success Rate")
        ax.set_xlabel("Success Rate")
        ax.set_ylabel("Frequency")
        return ax

    @staticmethod
    def boxplot_production_by_day(df, ax=None):
        if ax is None:
            fig, ax = plt.subplots(figsize=(8, 6))
        sns.boxplot(x="Day", y="Production Plan", data=df, ax=ax)
        ax.set_title("Production Plan by Day (boxplot)")
        ax.set_xlabel("Day")
        ax.set_ylabel("Production Plan")
        return ax

    @staticmethod
    def bar_actual_by_day(df, ax=None):
        if ax is None:
            fig, ax = plt.subplots(figsize=(8, 6))
        order = df["Day"].unique()
        agg = df.groupby("Day")["Actual Production"].mean().reindex(order)
        agg.plot(kind="bar", ax=ax, color="skyblue")
        ax.set_title("Mean Actual Production by Day")
        ax.set_xlabel("Day")
        ax.set_ylabel("Mean Actual Production")
        return ax

    @staticmethod
    def pairplot_numeric(df):
        numeric = df.select_dtypes("number")
        if numeric.shape[1] < 2:
            return None
        g = sns.pairplot(numeric)
        return g

    @staticmethod
    def residuals_plot(model, X_test, y_test, ax=None):
        if ax is None:
            fig, ax = plt.subplots(figsize=(8, 6))
        y_pred = model.predict(X_test)
        residuals = y_test - y_pred
        ax.scatter(y_pred, residuals, color="purple")
        ax.axhline(0, color="black", linestyle="--")
        ax.set_xlabel("Predicted Actual Production")
        ax.set_ylabel("Residuals")
        ax.set_title("Residuals vs Predicted")
        return ax

    @staticmethod
    def qq_plot(residuals, ax=None):
        if ax is None:
            fig, ax = plt.subplots(figsize=(6, 6))
        sm.qqplot(residuals, line="45", ax=ax)
        ax.set_title("QQ Plot of Residuals")
        return ax

    @staticmethod
    def violin_success_by_day(df, ax=None):
        if ax is None:
            fig, ax = plt.subplots(figsize=(8, 6))
        sns.violinplot(x="Day", y="Success Rate", data=df, ax=ax, inner="quartile")
        ax.set_title("Success Rate by Day (violin)")
        ax.set_xlabel("Day")
        ax.set_ylabel("Success Rate")
        return ax

    @staticmethod
    def heatmap_correlation(df, ax=None):
        if ax is None:
            fig, ax = plt.subplots(figsize=(8, 6))
        corr = df.select_dtypes("number").corr()
        sns.heatmap(corr, annot=True, fmt=".2f", cmap="coolwarm", ax=ax)
        ax.set_title("Correlation heatmap")
        return ax

    @staticmethod
    def cumulative_actual(df, ax=None):
        if ax is None:
            fig, ax = plt.subplots(figsize=(8, 6))
        df_sorted = df.copy()
        df_sorted = df_sorted.sort_index()
        if "Actual Production" not in df_sorted.columns:
            ax.set_title("No Actual Production column")
            return ax
        cum = df_sorted["Actual Production"].fillna(0).cumsum()
        ax.plot(cum.index, cum.values, marker="o")
        ax.set_title("Cumulative Actual Production")
        ax.set_xlabel("Row index")
        ax.set_ylabel("Cumulative Actual Production")
        return ax

    @staticmethod
    def stacked_plan_actual_by_day(df, ax=None):
        if ax is None:
            fig, ax = plt.subplots(figsize=(8, 6))
        agg = df.groupby("Day")[["Production Plan", "Actual Production"]].sum()
        agg = agg.reindex(df["Day"].unique())
        agg.plot(kind="bar", stacked=True, ax=ax)
        ax.set_title("Stacked Production Plan and Actual by Day")
        ax.set_xlabel("Day")
        ax.set_ylabel("Total")
        return ax


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Production Analyzer")
        self.geometry("1250x860")
        self.filepath = tk.StringVar()
        self.reader = None
        self.aggregator = None
        self.per_stuff = {}
        self.selected_key = None
        self.show_before_after = tk.BooleanVar(value=True)
        self._build_ui()

    def _build_ui(self):
        frm_top = ttk.Frame(self)
        frm_top.pack(fill="x", padx=8, pady=6)

        ttk.Label(frm_top, text="Excel file").pack(side="left")
        ttk.Entry(frm_top, textvariable=self.filepath, width=60).pack(side="left", padx=6)
        ttk.Button(frm_top, text="Browse", command=self.browse_file).pack(side="left", padx=4)
        ttk.Button(frm_top, text="Load & Aggregate", command=self.load_and_aggregate).pack(side="left", padx=4)

        chk = ttk.Checkbutton(frm_top, text="Show analysis before & after refill", variable=self.show_before_after)
        chk.pack(side="left", padx=8)

        frm_main = ttk.Frame(self)
        frm_main.pack(fill="both", expand=True, padx=8, pady=6)

        left = ttk.Frame(frm_main)
        left.pack(side="left", fill="y", padx=6, pady=6)

        ttk.Label(left, text="Stuff keys").pack(anchor="w")
        self.listbox = tk.Listbox(left, width=40, height=28)
        self.listbox.pack(fill="y")
        self.listbox.bind("<<ListboxSelect>>", self.on_select)

        btn_frame = ttk.Frame(left)
        btn_frame.pack(fill="x", pady=6)
        ttk.Button(btn_frame, text="Show DataFrame", command=self.show_dataframe).pack(side="left", padx=4)
        ttk.Button(btn_frame, text="Refill Missing", command=self.refill_missing).pack(side="left", padx=4)
        ttk.Button(btn_frame, text="Analyze", command=self.run_analysis).pack(side="left", padx=4)

        plot_frame = ttk.LabelFrame(left, text="Plot options")
        plot_frame.pack(fill="x", pady=6)
        self.plot_types = [
            "Scatter (Production Plan vs Actual Production)",
            "Histogram (Success Rate)",
            "Boxplot (Production Plan by Day)",
            "Bar (Mean Actual Production by Day)",
            "Pairplot (numeric columns)",
            "Residuals (regression)",
            "QQ plot (residuals)",
            "Violin (Success Rate by Day)",
            "Heatmap (correlation)",
            "Cumulative Actual",
            "Stacked Plan vs Actual by Day"
        ]
        self.plot_selector = ttk.Combobox(plot_frame, values=self.plot_types, state="readonly")
        self.plot_selector.current(0)
        self.plot_selector.pack(fill="x", padx=4, pady=4)
        ttk.Button(plot_frame, text="Show Plot", command=self.show_selected_plot).pack(fill="x", padx=4, pady=4)

        save_frame = ttk.Frame(left)
        save_frame.pack(fill="x", pady=6)
        ttk.Button(save_frame, text="Save Selected CSV", command=self.save_selected_csv).pack(side="left", padx=4)
        ttk.Button(save_frame, text="Save All CSVs", command=self.save_all_csvs).pack(side="left", padx=4)

        right = ttk.Frame(frm_main)
        right.pack(side="left", fill="both", expand=True, padx=6, pady=6)

        self.text = tk.Text(right, wrap="word", height=12)
        self.text.pack(fill="x", pady=4)

        plot_canvas_frame = ttk.Frame(right)
        plot_canvas_frame.pack(fill="both", expand=True)
        self.fig, self.ax = plt.subplots(figsize=(10, 6))
        self.canvas = FigureCanvasTkAgg(self.fig, master=plot_canvas_frame)
        self.canvas.get_tk_widget().pack(fill="both", expand=True)

    def browse_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
        if path:
            self.filepath.set(path)

    def load_and_aggregate(self):
        path = self.filepath.get().strip()
        if not path:
            messagebox.showwarning("No file", "Please choose an Excel file first.")
            return
        self.reader = ExcelReader(path)
        self.aggregator = DataAggregator(self.reader)
        sheet_range = [str(i) for i in range(27, 40) if i != 33]
        self.aggregator.aggregate(sheet_range)
        self.per_stuff = self.aggregator.to_per_stuff_dfs()
        self.listbox.delete(0, tk.END)
        for k in sorted(self.per_stuff.keys()):
            df = self.per_stuff[k]
            display_name = k if not df.empty else f"{k} (empty)"
            self.listbox.insert(tk.END, display_name)
        messagebox.showinfo("Done", f"Aggregated {len(self.per_stuff)} Stuff keys.")

    def _key_from_display(self, display_name):
        if display_name.endswith(" (empty)"):
            return display_name[: -len(" (empty)")]
        return display_name

    def on_select(self, event):
        sel = event.widget.curselection()
        if not sel:
            return
        idx = sel[0]
        display_key = event.widget.get(idx)
        key = self._key_from_display(display_key)
        self.selected_key = key

    def show_dataframe(self):
        if not self.selected_key:
            messagebox.showwarning("Select", "Select a Stuff key first.")
            return
        df = self.per_stuff.get(self.selected_key)
        top = tk.Toplevel(self)
        top.title(f"Data for {self.selected_key}")
        txt = tk.Text(top, wrap="none", width=140, height=40)
        txt.pack(fill="both", expand=True)
        if df is None:
            txt.insert("1.0", "No DataFrame found for this key.")
        elif df.empty:
            cols = ", ".join(df.columns.tolist())
            txt.insert("1.0", f"(empty DataFrame)\nColumns: {cols}\n\n")
            txt.insert("end", df.to_string(index=False))
        else:
            txt.insert("1.0", df.to_string(index=False))

    def _analyze_df(self, df, title_prefix=""):
        if df is None or df.empty:
            return f"{title_prefix}(empty DataFrame)\n", None
        analyzer = Analyzer(df)
        analyzer.clean_numeric()
        t_stat, p_value = analyzer.t_test_plan_vs_actual()
        corr = analyzer.correlation_matrix()
        reg_res = analyzer.linear_regression()
        text = f"{title_prefix}T-test (Production Plan vs Actual Production):\n  t = {t_stat}\n  p = {p_value}\n\n"
        text += "Correlation matrix:\n"
        text += corr.to_string() + "\n\n"
        if reg_res:
            text += f"Linear regression:\n  coef = {reg_res['coef']}\n  intercept = {reg_res['intercept']}\n  R2 = {reg_res['r2']}\n"
        else:
            text += "Linear regression: not enough data\n"
        return text, reg_res

    def refill_missing(self):
        if not self.selected_key:
            messagebox.showwarning("Select", "Select a Stuff key first.")
            return
        df = self.per_stuff.get(self.selected_key)
        if df is None:
            messagebox.showwarning("No data", "No DataFrame for selected key.")
            return

        show_both = bool(self.show_before_after.get())
        self.text.delete("1.0", tk.END)

        if show_both:
            before_text, before_reg = self._analyze_df(df, title_prefix="Before refill:\n")
            self.text.insert(tk.END, before_text + "\n")
            self.ax.clear()
            if before_reg:
                Xb = before_reg["X_test"].ravel()
                yb = before_reg["y_test"]
                Visualizer.scatter_with_regression(Xb, yb, model=before_reg["model"], title=f"Before refill - {self.selected_key}", xlabel="Production Plan", ylabel="Actual Production", ax=self.ax)
            else:
                df_plot = Analyzer(df).df.dropna(subset=["Production Plan", "Actual Production"])
                if not df_plot.empty:
                    self.ax.scatter(df_plot["Production Plan"], df_plot["Actual Production"], color="green")
                    self.ax.set_title(f"Before refill - {self.selected_key}")
                else:
                    self.ax.set_title(f"No data to plot before refill for {self.selected_key}")
            self.canvas.draw()

        analyzer = Analyzer(df)
        analyzer.clean_numeric()
        analyzer.refill_success_rate()
        filled = analyzer.refill_actual_from_success()
        self.per_stuff[self.selected_key] = analyzer.df

        if show_both:
            after_text, after_reg = self._analyze_df(analyzer.df, title_prefix="After refill:\n")
            self.text.insert(tk.END, f"Refill: rows filled = {filled}\n\n")
            self.text.insert(tk.END, after_text + "\n")
            self.ax.clear()
            if after_reg:
                Xa = after_reg["X_test"].ravel()
                ya = after_reg["y_test"]
                Visualizer.scatter_with_regression(Xa, ya, model=after_reg["model"], title=f"After refill - {self.selected_key}", xlabel="Production Plan", ylabel="Actual Production", ax=self.ax)
            else:
                df_plot = Analyzer(analyzer.df).df.dropna(subset=["Production Plan", "Actual Production"])
                if not df_plot.empty:
                    self.ax.scatter(df_plot["Production Plan"], df_plot["Actual Production"], color="green")
                    self.ax.set_title(f"After refill - {self.selected_key}")
                else:
                    self.ax.set_title(f"No data to plot after refill for {self.selected_key}")
            self.canvas.draw()
        else:
            messagebox.showinfo("Refill", f"Missing values refilled where possible. Rows filled: {filled}")
            self._refresh_listbox_entry(self.selected_key)

        self._refresh_listbox_entry(self.selected_key)

    def _refresh_listbox_entry(self, key):
        items = list(self.listbox.get(0, tk.END))
        for idx, display in enumerate(items):
            base = self._key_from_display(display)
            if base == key:
                df = self.per_stuff.get(key)
                new_display = key if df is not None and not df.empty else f"{key} (empty)"
                self.listbox.delete(idx)
                self.listbox.insert(idx, new_display)
                return

    def save_selected_csv(self):
        if not self.selected_key:
            messagebox.showwarning("Select", "Select a Stuff key first.")
            return
        df = self.per_stuff.get(self.selected_key)
        if df is None:
            messagebox.showwarning("No data", "No DataFrame for selected key.")
            return
        default_name = f"{self.selected_key}.csv"
        path = filedialog.asksaveasfilename(defaultextension=".csv", initialfile=default_name,
                                            filetypes=[("CSV files", "*.csv")])
        if not path:
            return
        try:
            df.to_csv(path, index=False, encoding="utf-8-sig")
            messagebox.showinfo("Saved", f"Saved {self.selected_key} to {path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save CSV: {e}")

    def save_all_csvs(self):
        if not self.per_stuff:
            messagebox.showwarning("No data", "No DataFrames to save.")
            return
        folder = filedialog.askdirectory()
        if not folder:
            return
        failed = []
        for key, df in self.per_stuff.items():
            safe_name = "".join(c if c.isalnum() or c in " _-." else "_" for c in str(key))
            path = os.path.join(folder, f"{safe_name}.csv")
            try:
                if df is None:
                    df_to_save = pd.DataFrame(columns=["Day", "Production Plan", "Actual Production", "Success Rate"])
                else:
                    df_to_save = df
                df_to_save.to_csv(path, index=False, encoding="utf-8-sig")
            except Exception as e:
                failed.append((key, str(e)))
        if not failed:
            messagebox.showinfo("Saved", f"Saved {len(self.per_stuff)} CSV files to {folder}")
        else:
            messagebox.showwarning("Partial success", f"Saved some files. Failures: {failed}")

    def run_analysis(self):
        if not self.selected_key:
            messagebox.showwarning("Select", "Select a Stuff key first.")
            return
        df = self.per_stuff.get(self.selected_key)
        if df is None or df.empty:
            self.text.delete("1.0", tk.END)
            self.text.insert(tk.END, f"Selected: {self.selected_key}\n\n(empty DataFrame)\n")
            self.ax.clear()
            self.ax.set_title(f"No data to plot for {self.selected_key}")
            self.canvas.draw()
            return

        text, reg_res = self._analyze_df(df, title_prefix=f"Analysis for {self.selected_key}:\n")
        self.text.delete("1.0", tk.END)
        self.text.insert(tk.END, text)

        self.ax.clear()
        if reg_res:
            X = reg_res["X_test"].ravel()
            y = reg_res["y_test"]
            Visualizer.scatter_with_regression(X, y, model=reg_res["model"], title=f"Regression for {self.selected_key}", xlabel="Production Plan", ylabel="Actual Production", ax=self.ax)
        else:
            df_plot = Analyzer(df).df.dropna(subset=["Production Plan", "Actual Production"])
            if not df_plot.empty:
                self.ax.scatter(df_plot["Production Plan"], df_plot["Actual Production"], color="green")
                self.ax.set_xlabel("Production Plan")
                self.ax.set_ylabel("Actual Production")
                self.ax.set_title(f"Production Plan vs Actual Production for {self.selected_key}")
            else:
                self.ax.set_title(f"No numeric data to plot for {self.selected_key}")
        self.canvas.draw()

    def _show_plot_data(self, df_used, title):
        top = tk.Toplevel(self)
        top.title(f"{title} - Data")
        txt = tk.Text(top, wrap="none", width=140, height=30)
        txt.pack(fill="both", expand=True)
        if df_used is None:
            txt.insert("1.0", "(no data)")
            return
        try:
            txt.insert("1.0", df_used.to_string(index=False))
        except Exception:
            txt.insert("1.0", df_used.to_csv(index=False))

    def show_selected_plot(self):
        if not self.selected_key:
            messagebox.showwarning("Select", "Select a Stuff key first.")
            return
        plot_choice = self.plot_selector.get()
        df = self.per_stuff.get(self.selected_key)
        if df is None or df.empty:
            messagebox.showinfo("No data", "Selected DataFrame is empty or missing.")
            return

        analyzer = Analyzer(df)
        analyzer.clean_numeric()
        self.ax.clear()

        df_used = df.copy()

        if plot_choice == "Scatter (Production Plan vs Actual Production)":
            df_plot = analyzer.df.dropna(subset=["Production Plan", "Actual Production"])
            df_used = df_plot.copy()
            if df_plot.empty:
                self.ax.set_title("No numeric data for scatter")
            else:
                self.ax.scatter(df_plot["Production Plan"], df_plot["Actual Production"], color="blue")
                self.ax.set_xlabel("Production Plan")
                self.ax.set_ylabel("Actual Production")
                self.ax.set_title(f"Scatter for {self.selected_key}")

        elif plot_choice == "Histogram (Success Rate)":
            df_used = analyzer.df[["Success Rate"]].copy()
            if "Success Rate" not in analyzer.df.columns or analyzer.df["Success Rate"].dropna().empty:
                self.ax.set_title("No Success Rate data")
            else:
                Visualizer.hist_success_rate(analyzer.df, ax=self.ax)

        elif plot_choice == "Boxplot (Production Plan by Day)":
            df_used = analyzer.df[["Day", "Production Plan"]].copy()
            if "Day" not in analyzer.df.columns or analyzer.df["Production Plan"].dropna().empty:
                self.ax.set_title("No data for boxplot")
            else:
                Visualizer.boxplot_production_by_day(analyzer.df, ax=self.ax)

        elif plot_choice == "Bar (Mean Actual Production by Day)":
            order = analyzer.df["Day"].unique() if "Day" in analyzer.df.columns else None
            agg = analyzer.df.groupby("Day")["Actual Production"].mean()
            df_used = agg.reset_index().rename(columns={"Actual Production": "Mean Actual Production"})
            if "Day" not in analyzer.df.columns or analyzer.df["Actual Production"].dropna().empty:
                self.ax.set_title("No data for bar plot")
            else:
                Visualizer.bar_actual_by_day(analyzer.df, ax=self.ax)

        elif plot_choice == "Pairplot (numeric columns)":
            numeric = analyzer.df.select_dtypes("number")
            df_used = numeric.copy()
            if numeric.shape[1] < 2:
                self.ax.set_title("Not enough numeric columns for pairplot")
            else:
                plt.close(self.fig)
                g = Visualizer.pairplot_numeric(analyzer.df)
                self._show_plot_data(df_used, "Pairplot data")
                return

        elif plot_choice == "Residuals (regression)":
            reg_res = analyzer.linear_regression()
            if not reg_res:
                self.ax.set_title("Not enough data for regression residuals")
                df_used = pd.DataFrame(columns=["X_test", "y_test"])
            else:
                y_pred = reg_res["model"].predict(reg_res["X_test"])
                residuals = reg_res["y_test"] - y_pred
                df_used = pd.DataFrame(reg_res["X_test"], columns=[f"feature_{i}" for i in range(reg_res["X_test"].shape[1])])
                df_used["y_test"] = reg_res["y_test"]
                df_used["y_pred"] = y_pred
                df_used["residual"] = residuals
                Visualizer.residuals_plot(reg_res["model"], reg_res["X_test"], reg_res["y_test"], ax=self.ax)

        elif plot_choice == "QQ plot (residuals)":
            reg_res = analyzer.linear_regression()
            if not reg_res:
                self.ax.set_title("Not enough data for QQ plot")
                df_used = pd.DataFrame(columns=["X_test", "y_test"])
            else:
                y_pred = reg_res["model"].predict(reg_res["X_test"])
                residuals = reg_res["y_test"] - y_pred
                df_used = pd.DataFrame({"residuals": residuals})
                Visualizer.qq_plot(residuals, ax=self.ax)

        elif plot_choice == "Violin (Success Rate by Day)":
            df_used = analyzer.df[["Day", "Success Rate"]].copy()
            if "Success Rate" not in analyzer.df.columns or analyzer.df["Success Rate"].dropna().empty:
                self.ax.set_title("No Success Rate data for violin")
            else:
                Visualizer.violin_success_by_day(analyzer.df, ax=self.ax)

        elif plot_choice == "Heatmap (correlation)":
            numeric = analyzer.df.select_dtypes("number")
            df_used = numeric.copy()
            if numeric.shape[1] < 2:
                self.ax.set_title("Not enough numeric columns for heatmap")
            else:
                Visualizer.heatmap_correlation(analyzer.df, ax=self.ax)

        elif plot_choice == "Cumulative Actual":
            df_sorted = analyzer.df.copy().sort_index()
            df_used = df_sorted[["Actual Production"]].copy()
            if "Actual Production" not in df_sorted.columns or df_sorted["Actual Production"].dropna().empty:
                self.ax.set_title("No Actual Production data for cumulative plot")
            else:
                Visualizer.cumulative_actual(analyzer.df, ax=self.ax)

        elif plot_choice == "Stacked Plan vs Actual by Day":
            agg = analyzer.df.groupby("Day")[["Production Plan", "Actual Production"]].sum().reset_index()
            df_used = agg.copy()
            if "Day" not in analyzer.df.columns or (analyzer.df[["Production Plan", "Actual Production"]].dropna().empty):
                self.ax.set_title("Not enough data for stacked bar")
            else:
                Visualizer.stacked_plan_actual_by_day(analyzer.df, ax=self.ax)

        self.canvas.draw()
        self._show_plot_data(df_used, f"{plot_choice} data")


if __name__ == "__main__":
    app = App()
    app.mainloop()
