import sys
import unicodedata
from collections import defaultdict
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from scipy import stats
from sklearn.linear_model import LinearRegression
from sklearn.model_selection import train_test_split
import matplotlib.pyplot as plt
import seaborn as sns

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass


class ExcelReader:
    def __init__(self, path, max_col_limit=22):
        self.path = path
        self.max_col_limit = max_col_limit

    @staticmethod
    def _normalize_cols(cols):
        return [unicodedata.normalize("NFKC", str(c)).strip() for c in cols]

    def _safe_usecols_for_sheet(self, sheet_name):
        try:
            wb = load_workbook(self.path, read_only=True)
            ws = wb[str(sheet_name)]
            max_col = ws.max_column
            last_col_letter = get_column_letter(min(max_col, self.max_col_limit))
            return f"A:{last_col_letter}"
        except Exception:
            return None

    def read(self, sheet_name, usecols=None):
        if usecols is None:
            usecols = self._safe_usecols_for_sheet(sheet_name)
        try:
            if usecols:
                df = pd.read_excel(self.path, sheet_name=str(sheet_name), usecols=usecols, engine="openpyxl")
            else:
                df = pd.read_excel(self.path, sheet_name=str(sheet_name), engine="openpyxl")
            df.columns = self._normalize_cols(df.columns)
            return df
        except Exception:
            try:
                df = pd.read_excel(self.path, sheet_name=str(sheet_name), engine="openpyxl")
                df.columns = self._normalize_cols(df.columns)
                return df
            except Exception as e2:
                print(f"Failed to read sheet {sheet_name}: {e2}")
                return None


class DataAggregator:
    def __init__(self, reader, days=None):
        self.reader = reader
        self.days = days or ["Saturday", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
        self.acc = defaultdict(lambda: defaultdict(list))

    def collect_production_names(self, sheet_range):
        names = set()
        for sheet in sheet_range:
            df = self.reader.read(sheet_name=str(sheet), usecols="A")
            if df is None:
                continue
            if "Stuff" in df.columns:
                names.update(df["Stuff"].dropna().astype(str).tolist())
        return names

    def aggregate(self, sheet_range):
        for sheet in sheet_range:
            df = self.reader.read(sheet_name=str(sheet), usecols="A:V")
            if df is None:
                continue
            if "Stuff" not in df.columns:
                continue
            for _, row in df.iterrows():
                stuff = str(row["Stuff"]).strip()
                if not stuff or stuff.lower() == "nan":
                    continue
                row_list = list(row)
                for i, day in enumerate(self.days):
                    start = 1 + i * 3
                    block = row_list[start:start + 3]
                    if all(pd.isna(x) for x in block):
                        continue
                    self.acc[stuff][day].append(block)

    def to_per_stuff_dfs(self):
        per_stuff = {}
        for stuff, daydict in self.acc.items():
            frames = []
            for day in self.days:
                blocks = daydict.get(day, [])
                if not blocks:
                    continue
                df_day = pd.DataFrame(blocks, columns=["Production Plan", "Actual Production", "Success Rate"])
                df_day["Day"] = day
                df_day = df_day[["Day", "Production Plan", "Actual Production", "Success Rate"]]
                frames.append(df_day)
            if frames:
                per_stuff[stuff] = pd.concat(frames, ignore_index=True)
            else:
                per_stuff[stuff] = pd.DataFrame(columns=["Day", "Production Plan", "Actual Production", "Success Rate"])
        return per_stuff


class Analyzer:
    def __init__(self, df):
        self.df = df.copy()

    def clean_numeric(self):
        for col in ["Production Plan", "Actual Production", "Success Rate"]:
            if col in self.df.columns:
                self.df[col] = pd.to_numeric(self.df[col], errors="coerce")

    def t_test_plan_vs_actual(self, drop_zeros=True):
        df = self.df
        if drop_zeros:
            df = df[df["Actual Production"] != 0]
        a = df["Production Plan"].dropna()
        b = df["Actual Production"].dropna()
        if len(a) < 2 or len(b) < 2:
            return None, None
        t_stat, p_value = stats.ttest_ind(a, b, nan_policy="omit")
        return t_stat, p_value

    def correlation_matrix(self, method="pearson"):
        return self.df.select_dtypes("number").corr(method=method)

    def linear_regression(self, feature="Production Plan", target="Actual Production", test_size=0.2, random_state=42):
        X = self.df[[feature]].dropna()
        y = self.df.loc[X.index, target].dropna()
        X = X.loc[y.index]
        if len(X) < 2:
            return None
        X_train, X_test, y_train, y_test = train_test_split(X.values, y.values, test_size=test_size, random_state=random_state)
        model = LinearRegression().fit(X_train, y_train)
        y_pred = model.predict(X_test)
        return {
            "model": model,
            "X_test": X_test,
            "y_test": y_test,
            "y_pred": y_pred,
            "r2": model.score(X_test, y_test),
            "coef": model.coef_,
            "intercept": model.intercept_,
        }

    def refill_success_rate(self):
        df = self.df
        if "Success Rate" in df.columns and "Production Plan" in df.columns and "Actual Production" in df.columns:
            mask = df["Success Rate"].isna() & df["Production Plan"].notna() & (df["Production Plan"] != 0)
            df.loc[mask, "Success Rate"] = (df.loc[mask, "Actual Production"] / df.loc[mask, "Production Plan"]) * 100
            self.df = df

    def refill_actual_from_success(self):
        df = self.df
        if "Actual Production" in df.columns and "Production Plan" in df.columns and "Success Rate" in df.columns:
            mask = df["Actual Production"].isna() & df["Success Rate"].notna()
            df.loc[mask, "Actual Production"] = df.loc[mask, "Production Plan"] * (df.loc[mask, "Success Rate"] / 100)
            self.df = df


class Visualizer:
    @staticmethod
    def scatter_with_regression(X, y, model=None, title=None, xlabel=None, ylabel=None):
        plt.figure(figsize=(8, 6))
        plt.scatter(X, y, color="blue", label="Data points")
        if model is not None:
            xs = np.linspace(np.nanmin(X), np.nanmax(X), 100).reshape(-1, 1)
            plt.plot(xs, model.predict(xs), color="red", linewidth=2, label="Regression line")
        plt.xlabel(xlabel or "X")
        plt.ylabel(ylabel or "Y")
        if title:
            plt.title(title)
        plt.legend()
        plt.show()

    @staticmethod
    def catplot_by_day(df, y_col, title=None):
        sns.catplot(x="Day", y=y_col, data=df, kind="bar", height=5, aspect=1.5)
        plt.title(title or f"{y_col} by Day")
        plt.show()

    @staticmethod
    def hist_success_rate(df, title=None):
        plt.figure(figsize=(8, 6))
        sns.histplot(df["Success Rate"].dropna(), bins=20, kde=True)
        plt.title(title or "Distribution of Success Rate")
        plt.xlabel("Success Rate")
        plt.ylabel("Frequency")
        plt.show()


def main():
    FILE = r"D:\UT_3\Data_Analysis\DataSet\New Microsoft Excel Worksheet.xlsx"
    reader = ExcelReader(FILE)
    aggregator = DataAggregator(reader, days=["Saturday", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday"])

    sheet_range = [str(i) for i in range(27, 40) if i != 33]
    aggregator.aggregate(sheet_range)
    per_stuff = aggregator.to_per_stuff_dfs()

    if not per_stuff:
        print("No data aggregated.")
        return

    example_key = next(iter(per_stuff))
    df_example = per_stuff[example_key]

    analyzer = Analyzer(df_example)
    analyzer.clean_numeric()

    t_stat, p_value = analyzer.t_test_plan_vs_actual()
    print(f"T-test for {example_key}: t={t_stat}, p={p_value}")

    corr = analyzer.correlation_matrix()
    print("Correlation matrix:\n", corr)

    reg_res = analyzer.linear_regression()
    if reg_res:
        print("Linear regression coef:", reg_res["coef"], "intercept:", reg_res["intercept"], "R2:", reg_res["r2"])
        Visualizer.scatter_with_regression(reg_res["X_test"].ravel(), reg_res["y_test"], model=reg_res["model"],
                                          title=f"Regression for {example_key}", xlabel="Production Plan", ylabel="Actual Production")

    analyzer.refill_success_rate()
    analyzer.refill_actual_from_success()
    print("Updated DataFrame:\n", analyzer.df)

    Visualizer.catplot_by_day(analyzer.df, "Actual Production", title=f"Actual Production by Day for {example_key}")
    Visualizer.hist_success_rate(analyzer.df, title=f"Success Rate Distribution for {example_key}")


if __name__ == "__main__":
    main()