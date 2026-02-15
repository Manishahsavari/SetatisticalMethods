from __future__ import annotations

import os
import sys
import re
import unicodedata
import base64
import subprocess
import tempfile
import shutil
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy import stats
from sklearn.linear_model import LinearRegression
from sklearn.model_selection import train_test_split

import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext

import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk

import plotly.express as px
import plotly.graph_objects as go

# Check for kaleido (image export for plotly)
_HAS_KALEIDO = False
try:
    import plotly.io as pio
    _HAS_KALEIDO = True
except ImportError:
    pass

# RTL and webview support
_HAS_RTL_LIBS = False
_HAS_WEBVIEW = False

try:
    import arabic_reshaper
    from bidi.algorithm import get_display
    _HAS_RTL_LIBS = True
except ImportError:
    pass

try:
    import webview
    _HAS_WEBVIEW = True
except ImportError:
    pass

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

# ==================== Helper Functions ====================
def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return unicodedata.normalize("NFKC", str(value)).strip()

def to_numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")

def has_persian_or_arabic(text: str) -> bool:
    return bool(re.search(r"[\u0600-\u06FF\u0750-\u077F\u08A0-\u08FF]", text or ""))

def fix_rtl_display(text: object) -> str:
    if text is None or (isinstance(text, float) and np.isnan(text)):
        return ""
    txt = normalize_text(text)
    if not txt:
        return ""
    if _HAS_RTL_LIBS and has_persian_or_arabic(txt):
        try:
            reshaped = arabic_reshaper.reshape(txt)
            return get_display(reshaped)
        except Exception:
            return txt
    return txt

def fix_rtl_for_html(text: object) -> str:
    return normalize_text(text) if text is not None else ""

def get_script_dir() -> str:
    try:
        return os.path.dirname(os.path.abspath(__file__))
    except Exception:
        return os.getcwd()

def find_local_font() -> Optional[str]:
    dir_ = get_script_dir()
    candidates = [
        "Vazirmatn-Regular.ttf", "vazirmatn-regular.ttf",
        "Vazir-Regular.ttf", "vazir-regular.ttf",
        "Vazir.ttf", "vazir.ttf",
        "IRANSans.ttf", "iransans.ttf",
    ]
    for name in candidates:
        path = os.path.join(dir_, name)
        if os.path.exists(path):
            return os.path.abspath(path)
    return None

def font_to_data_url(font_path: str) -> str:
    with open(font_path, "rb") as f:
        b64 = base64.b64encode(f.read()).decode("ascii")
    return f"data:font/ttf;base64,{b64}"

def get_embedded_font_css(font_family: str = "Vazirmatn") -> str:
    font_path = find_local_font()
    if not font_path:
        return ""
    try:
        data_url = font_to_data_url(font_path)
        return f"""@font-face {{
            font-family: '{font_family}';
            src: url('{data_url}') format('truetype');
        }}"""
    except Exception:
        return ""

def open_plotly_in_app(html_path: str, title: str = "Dashboard") -> None:
    if _HAS_WEBVIEW:
        try:
            webview.create_window(title=title, url=f"file://{os.path.abspath(html_path)}",
                                  width=1200, height=820)
            webview.start(gui="tkinter", debug=False)
            return
        except Exception:
            pass
    import webbrowser
    webbrowser.open_new_tab(f"file://{os.path.abspath(html_path)}")

def configure_matplotlib_persian_font() -> None:
    try:
        from matplotlib import font_manager
        local_font = find_local_font()
        if local_font:
            try:
                font_manager.fontManager.addfont(local_font)
                prop = font_manager.FontProperties(fname=local_font)
                plt.rcParams["font.family"] = prop.get_name()
            except Exception:
                pass
    except Exception:
        pass
    preferred = ["Vazirmatn", "Vazir", "IRANSans", "DejaVu Sans"]
    try:
        from matplotlib.font_manager import fontManager
        available = {f.name for f in fontManager.ttflist}
        for f in preferred:
            if f in available:
                plt.rcParams["font.family"] = f
                break
        plt.rcParams["axes.unicode_minus"] = False
    except Exception:
        plt.rcParams["axes.unicode_minus"] = False

DAYS = ["Saturday", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

# ==================== Data Processing Classes ====================
class ExcelReader:
    def __init__(self, path: str):
        self.path = path

    def list_sheets(self) -> List[str]:
        wb = load_workbook(self.path, read_only=True, data_only=True)
        return [str(s) for s in wb.sheetnames]

    def read(self, sheet_name: str, usecols: Optional[str] = None) -> pd.DataFrame:
        df = pd.read_excel(self.path, sheet_name=str(sheet_name), usecols=usecols, engine="openpyxl")
        df.columns = [normalize_text(c) for c in df.columns]
        return df

@dataclass
class DayBlockSpec:
    day: str
    plan_col: int
    actual_col: int
    success_col: int

class DatasetParser:
    def __init__(self, days: List[str] = DAYS):
        self.days = days

    def detect_day_blocks(self, cols: List[str]) -> List[DayBlockSpec]:
        col_map = {i: normalize_text(c) for i, c in enumerate(cols)}
        blocks = []
        for i, name in col_map.items():
            if name in self.days and i + 2 < len(cols):
                blocks.append(DayBlockSpec(day=name, plan_col=i, actual_col=i+1, success_col=i+2))
        day_to_block = {b.day: b for b in blocks}
        return [day_to_block[d] for d in self.days if d in day_to_block]

    def parse_sheet(self, df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
        if df is None or df.empty:
            return pd.DataFrame(columns=["Stuff", "Sheet", "Day", "Production Plan", "Actual Production", "Success Rate"])

        stuff_idx = None
        for i, col in enumerate(df.columns):
            if normalize_text(col).lower() == "stuff":
                stuff_idx = i
                break
        if stuff_idx is None:
            stuff_idx = 0

        blocks = self.detect_day_blocks(list(df.columns))
        if not blocks:
            blocks = []
            base = stuff_idx + 1
            for k, day in enumerate(self.days):
                idx = base + 3*k
                if idx + 2 < len(df.columns):
                    blocks.append(DayBlockSpec(day=day, plan_col=idx, actual_col=idx+1, success_col=idx+2))

        rows = []
        for _, row in df.iterrows():
            stuff = normalize_text(row.iloc[stuff_idx])
            if not stuff or stuff.lower() == "nan":
                continue
            for blk in blocks:
                plan = row.iloc[blk.plan_col] if blk.plan_col < len(row) else np.nan
                actual = row.iloc[blk.actual_col] if blk.actual_col < len(row) else np.nan
                success = row.iloc[blk.success_col] if blk.success_col < len(row) else np.nan
                if pd.isna(plan) and pd.isna(actual) and pd.isna(success):
                    continue
                rows.append([stuff, str(sheet_name), blk.day, plan, actual, success])

        result = pd.DataFrame(rows, columns=["Stuff", "Sheet", "Day", "Production Plan", "Actual Production", "Success Rate"])
        for col in ["Production Plan", "Actual Production", "Success Rate"]:
            result[col] = to_numeric(result[col])
        return result

class DataRepository:
    def __init__(self, reader: ExcelReader, parser: DatasetParser):
        self.reader = reader
        self.parser = parser
        self.df_all = pd.DataFrame()
        self.per_stuff: Dict[str, pd.DataFrame] = {}

    def load(self, sheet_names: List[str]) -> None:
        frames = []
        for sheet in sheet_names:
            try:
                df = self.reader.read(sheet)
                frames.append(self.parser.parse_sheet(df, sheet))
            except Exception:
                continue
        if frames:
            self.df_all = pd.concat(frames, ignore_index=True)
        else:
            self.df_all = pd.DataFrame(columns=["Stuff", "Sheet", "Day", "Production Plan", "Actual Production", "Success Rate"])
        self.per_stuff = {k: g.reset_index(drop=True) for k, g in self.df_all.groupby("Stuff")}

class Analyzer:
    def __init__(self, df: pd.DataFrame):
        self.df = df.copy()

    def clean(self) -> None:
        for col in ["Production Plan", "Actual Production", "Success Rate"]:
            if col in self.df.columns:
                self.df[col] = to_numeric(self.df[col])

    def refill_success_rate(self) -> int:
        self.clean()
        if not {"Success Rate", "Production Plan", "Actual Production"}.issubset(self.df.columns):
            return 0
        mask = (self.df["Success Rate"].isna() &
                self.df["Production Plan"].notna() &
                (self.df["Production Plan"] != 0) &
                self.df["Actual Production"].notna())
        filled = int(mask.sum())
        self.df.loc[mask, "Success Rate"] = (self.df.loc[mask, "Actual Production"] / self.df.loc[mask, "Production Plan"]) * 100.0
        return filled

    def refill_actual(self, use_success_rate: bool = True, min_train: int = 10) -> int:
        self.clean()
        if not {"Production Plan", "Actual Production"}.issubset(self.df.columns):
            return 0

        target_mask = (self.df["Production Plan"].notna() &
                       (self.df["Production Plan"] != 0) &
                       (self.df["Actual Production"].isna() | (self.df["Actual Production"] == 0)))

        features = ["Production Plan"]
        if use_success_rate and "Success Rate" in self.df.columns:
            features.append("Success Rate")

        train_mask = (self.df["Actual Production"].notna() &
                      (self.df["Actual Production"] != 0) &
                      self.df[features].notna().all(axis=1))
        n_train = int(train_mask.sum())
        filled = 0

        if n_train >= min_train:
            X_train = self.df.loc[train_mask, features].values
            y_train = self.df.loc[train_mask, "Actual Production"].values
            model = LinearRegression().fit(X_train, y_train)

            pred_mask = target_mask & self.df[features].notna().all(axis=1)
            if pred_mask.any():
                X_pred = self.df.loc[pred_mask, features].values
                y_pred = model.predict(X_pred)
                y_pred = np.where(np.isfinite(y_pred), np.maximum(0, y_pred), np.nan)
                self.df.loc[pred_mask, "Actual Production"] = y_pred
                filled = int(pred_mask.sum())
        else:
            if "Success Rate" in self.df.columns:
                fallback = target_mask & self.df["Success Rate"].notna()
                if fallback.any():
                    self.df.loc[fallback, "Actual Production"] = (self.df.loc[fallback, "Production Plan"] *
                                                                   (self.df.loc[fallback, "Success Rate"] / 100.0))
                    filled = int(fallback.sum())
        return filled

    def t_test_plan_vs_actual(self, drop_zeros: bool = True) -> Tuple[Optional[float], Optional[float], int, int]:
        self.clean()
        df = self.df.copy()
        if drop_zeros:
            df = df[df["Actual Production"].notna() & (df["Actual Production"] != 0)]
        plan = df["Production Plan"].dropna()
        actual = df["Actual Production"].dropna()
        if len(plan) < 2 or len(actual) < 2:
            return None, None, len(plan), len(actual)
        t, p = stats.ttest_ind(plan, actual, nan_policy="omit", equal_var=False)
        return float(t), float(p), len(plan), len(actual)

    def correlation(self, method: str = "pearson") -> pd.DataFrame:
        self.clean()
        return self.df[["Production Plan", "Actual Production", "Success Rate"]].corr(method=method)

    def regression_plan_to_actual(self, test_size: float = 0.2, random_state: int = 42) -> Optional[dict]:
        self.clean()
        data = self.df.dropna(subset=["Production Plan", "Actual Production"])
        if len(data) < 5:
            return None
        X = data[["Production Plan"]].values
        y = data["Actual Production"].values
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=test_size, random_state=random_state)
        model = LinearRegression().fit(X_train, y_train)
        r2 = model.score(X_test, y_test)
        return {"coef": float(model.coef_[0]), "intercept": float(model.intercept_), "r2": float(r2)}

# ==================== Visualization Classes ====================
class MatplotlibViz:
    @staticmethod
    def apply_modern_style():
        plt.rcParams.update({
            "axes.grid": True, "grid.alpha": 0.25, "axes.titleweight": "bold",
            "axes.titlesize": 12, "axes.labelsize": 10, "figure.dpi": 110,
        })

    @staticmethod
    def scatter_plan_actual(ax, df: pd.DataFrame, title: str):
        data = df.dropna(subset=["Production Plan", "Actual Production"])
        ax.clear()
        if data.empty:
            ax.set_title("No numeric data")
            return
        ax.scatter(data["Production Plan"], data["Actual Production"], alpha=0.8)
        ax.set_xlabel("Production Plan")
        ax.set_ylabel("Actual Production")
        ax.set_title(title)

class PlotlyViz:
    @staticmethod
    def interactive_dashboard(df: pd.DataFrame, title: str) -> str:
        title_disp = fix_rtl_for_html(title)
        data = df.copy()
        font_family = "Vazirmatn, Vazir, IRANSans, 'DejaVu Sans', Arial, sans-serif"

        fig1 = px.scatter(data, x="Production Plan", y="Actual Production", color="Day",
                          hover_data=["Sheet"], title=f"{title_disp} — Plan vs Actual")
        fig1.update_traces(marker=dict(size=9, opacity=0.85))
        fig1.update_layout(font=dict(size=14), title_font_size=20, font_family=font_family)

        fig2 = px.histogram(data, x="Success Rate", nbins=30, title=f"{title_disp} — Success Rate Distribution")
        fig2.update_layout(font=dict(size=14), title_font_size=20, font_family=font_family)

        agg = data.groupby("Day", as_index=False)[["Production Plan", "Actual Production"]].sum()
        agg["Day"] = pd.Categorical(agg["Day"], categories=DAYS, ordered=True)
        agg = agg.sort_values("Day")
        fig3 = go.Figure()
        fig3.add_bar(name="Production Plan", x=agg["Day"], y=agg["Production Plan"])
        fig3.add_bar(name="Actual Production", x=agg["Day"], y=agg["Actual Production"])
        fig3.update_layout(barmode="group", title=f"{title_disp} — Totals by Day",
                           font=dict(size=14), title_font_size=20, font_family=font_family)

        html = f"""
        <html><head><meta charset="utf-8"><title>{title_disp}</title>
        <style>{get_embedded_font_css('Vazirmatn')}
        body {{ font-family: {font_family}; margin:14px; background:#fff; color:#333; }}
        .grid {{ display:grid; grid-template-columns:1fr; gap:18px; }}
        .card {{ background:#f5f5f5; border:1px solid #ddd; border-radius:14px; padding:10px; }}
        h2 {{ margin:6px 0 0 4px; color:#58013B; }}
        .hint {{ margin:6px 0 14px 4px; color:#666; }}
        </style></head><body>
        <h2>{title_disp} — Interactive Dashboard</h2>
        <div class="grid">
            <div class="card">{fig1.to_html(include_plotlyjs="cdn", full_html=False)}</div>
            <div class="card">{fig2.to_html(include_plotlyjs=False, full_html=False)}</div>
            <div class="card">{fig3.to_html(include_plotlyjs=False, full_html=False)}</div>
        </div></body></html>
        """
        out_dir = os.path.join(os.path.expanduser("~"), ".production_analyzer")
        os.makedirs(out_dir, exist_ok=True)
        safe_name = re.sub(r"[^a-zA-Z0-9_-]+", "_", title)[:80]
        out_path = os.path.join(out_dir, f"{safe_name}_dashboard.html")
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(html)
        return out_path

    @staticmethod
    def static_dashboard_images(df: pd.DataFrame, stuff_name: str, out_dir: str) -> List[str]:
        if not _HAS_KALEIDO:
            raise ImportError("kaleido not installed. Run: pip install kaleido")
        import plotly.io as pio
        images = []
        data = df.copy()
        safe = re.sub(r"[^a-zA-Z0-9_-]+", "_", stuff_name)[:50]

        fig1 = px.scatter(data, x="Production Plan", y="Actual Production", color="Day", title="Plan vs Actual")
        path1 = os.path.join(out_dir, f"{safe}_scatter.png")
        pio.write_image(fig1, path1, format='png', width=800, height=500)
        images.append(path1)

        fig2 = px.histogram(data, x="Success Rate", nbins=30, title="Success Rate Distribution")
        path2 = os.path.join(out_dir, f"{safe}_hist.png")
        pio.write_image(fig2, path2, format='png', width=800, height=500)
        images.append(path2)

        agg = data.groupby("Day", as_index=False)[["Production Plan", "Actual Production"]].sum()
        agg["Day"] = pd.Categorical(agg["Day"], categories=DAYS, ordered=True)
        agg = agg.sort_values("Day")
        fig3 = px.bar(agg, x="Day", y=["Production Plan", "Actual Production"], barmode="group", title="Totals by Day")
        path3 = os.path.join(out_dir, f"{safe}_bar.png")
        pio.write_image(fig3, path3, format='png', width=800, height=500)
        images.append(path3)

        return images

class ReportViz:
    @staticmethod
    def _p_class(p: Optional[float]) -> str:
        if p is None or np.isnan(p) or np.isinf(p):
            return "na"
        return "good" if p < 0.05 else "bad"

    @staticmethod
    def _format(x, nd=6) -> str:
        if x is None:
            return "—"
        try:
            if isinstance(x, (float, np.floating)):
                if np.isnan(x) or np.isinf(x):
                    return "—"
                return f"{float(x):.{nd}f}"
            if isinstance(x, (int, np.integer)):
                return str(int(x))
        except:
            pass
        return str(x)

    @staticmethod
    def to_html_report(stuff: str, before: dict, after: dict, filled_sr: int, filled_act: int,
                       corr_before: pd.DataFrame, corr_after: pd.DataFrame,
                       reg_before: Optional[dict], reg_after: Optional[dict]) -> str:
        disp = fix_rtl_for_html(stuff)
        font = "Vazirmatn, Vazir, IRANSans, 'DejaVu Sans', Arial, sans-serif"
        cb = corr_before.round(6).to_html(index=True, border=0)
        ca = corr_after.round(6).to_html(index=True, border=0)
        t0, p0 = before.get("t"), before.get("p")
        t1, p1 = after.get("t"), after.get("p")

        def reg_block(reg):
            if not reg:
                return "<div class='muted'>Not enough data</div>"
            return f"""
            <div class='kpi-row'>
              <div class='kpi'><div class='k'>coef</div><div class='v'>{ReportViz._format(reg['coef'],6)}</div></div>
              <div class='kpi'><div class='k'>intercept</div><div class='v'>{ReportViz._format(reg['intercept'],6)}</div></div>
              <div class='kpi'><div class='k'>R²</div><div class='v'>{ReportViz._format(reg['r2'],6)}</div></div>
            </div>"""

        return f"""
        <html><head><meta charset='utf-8'><title>{disp} — Report</title>
        <style>
        @page {{ size: A4 landscape; margin: 1.5cm; }}
        :root {{ --bg:#fff; --card:#f5f5f5; --border:#ddd; --text:#333; --muted:#666; --good:#4caf50; --bad:#f44336; }}
        body {{ margin:16px; background:var(--bg); color:var(--text); font-family:{font}; }}
        .wrap {{ max-width:1100px; margin:0 auto; }}
        .header {{ display:flex; justify-content:space-between; margin-bottom:14px; }}
        h1 {{ font-size:22px; margin:0; color:#58013B; }}
        .sub {{ color:var(--muted); font-size:13px; }}
        .grid {{ display:grid; grid-template-columns:1fr; gap:14px; }}
        .card {{ background:var(--card); border:1px solid var(--border); border-radius:14px; padding:14px; }}
        h2 {{ font-size:16px; margin:0 0 10px; color:#58013B; }}
        .kpi-row {{ display:grid; grid-template-columns:repeat(3,1fr); gap:10px; }}
        .kpi {{ border:1px solid var(--border); border-radius:12px; padding:10px; background:#fff; }}
        .kpi .k {{ color:var(--muted); font-size:12px; }}
        .kpi .v {{ font-size:16px; font-weight:700; }}
        .muted {{ color:var(--muted); }}
        .good {{ color:var(--good); font-weight:800; }}
        .bad {{ color:var(--bad); font-weight:800; }}
        .pill {{ display:inline-block; padding:4px 10px; border-radius:999px; border:1px solid var(--border); background:#fff; font-size:12px; }}
        table {{ width:100%; border-collapse:collapse; margin-top:6px; font-size:13px; }}
        th, td {{ border:1px solid var(--border); padding:8px; text-align:center; background:#fff; }}
        th {{ background:#58013B; color:white; }}
        .two {{ display:grid; grid-template-columns:1fr 1fr; gap:14px; }}
        </style></head><body>
        <div class='wrap'>
            <div class='header'>
                <div><h1>{disp}</h1><div class='sub'>Before/After Refill • Plan vs Actual • Correlation • Regression</div></div>
                <div class='pill'>Filled SR: <b>{filled_sr}</b> • Filled Actual: <b>{filled_act}</b></div>
            </div>
            <div class='grid'>
                <div class='two'>
                    <div class='card'><h2>Before — T-test</h2>{ReportViz._format_kpi(t0,p0)}</div>
                    <div class='card'><h2>After — T-test</h2>{ReportViz._format_kpi(t1,p1)}</div>
                </div>
                <div class='two'>
                    <div class='card'><h2>Correlation — Before</h2>{cb}</div>
                    <div class='card'><h2>Correlation — After</h2>{ca}</div>
                </div>
                <div class='two'>
                    <div class='card'><h2>Regression — Before</h2>{reg_block(reg_before)}</div>
                    <div class='card'><h2>Regression — After</h2>{reg_block(reg_after)}</div>
                </div>
                <div class='card'><h2>Notes</h2><div class='muted'>• p<0.05 → significant difference (Welch).<br/>• Correlation strength near ±1.<br/>• R²: variance explained by plan.</div></div>
            </div>
        </div></body></html>
        """

    @staticmethod
    def _format_kpi(t, p):
        interp = "Significant" if (p is not None and p < 0.05) else "Not significant"
        pclass = ReportViz._p_class(p)
        return f"""
        <div class='kpi-row'>
            <div class='kpi'><div class='k'>t</div><div class='v'>{ReportViz._format(t,6)}</div></div>
            <div class='kpi'><div class='k'>p-value</div><div class='v {pclass}'>{ReportViz._format(p,6)}</div></div>
            <div class='kpi'><div class='k'>Interpretation</div><div class='v muted'>{interp}</div></div>
        </div>"""

    @staticmethod
    def save_report(html: str, title: str) -> str:
        out_dir = os.path.join(os.path.expanduser("~"), ".production_analyzer")
        os.makedirs(out_dir, exist_ok=True)
        safe = re.sub(r"[^a-zA-Z0-9_-]+", "_", title)[:80]
        out_path = os.path.join(out_dir, f"{safe}_report.html")
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(html)
        return out_path

# ==================== PDF Exporter (using weasyprint.exe) ====================
class PDFExporter:
    def __init__(self, weasyprint_exe_path: str = None):
        self.weasyprint_exe_path = weasyprint_exe_path
        self.temp_dir = os.path.join(os.path.expanduser("~"), ".production_analyzer", "temp_pdf")
        os.makedirs(self.temp_dir, exist_ok=True)

    def set_weasyprint_exe(self, path: str) -> bool:
        if os.path.exists(path) and path.lower().endswith('.exe'):
            self.weasyprint_exe_path = path
            return True
        return False

    def html_to_pdf(self, html_path: str, pdf_path: str) -> Tuple[bool, str]:
        if not self.weasyprint_exe_path or not os.path.exists(self.weasyprint_exe_path):
            return False, "weasyprint.exe path not set or missing."
        try:
            result = subprocess.run([self.weasyprint_exe_path, html_path, pdf_path],
                                     capture_output=True, text=True, timeout=60)
            if result.returncode == 0 and os.path.exists(pdf_path):
                return True, "PDF created."
            return False, f"weasyprint error: {result.stderr}"
        except subprocess.TimeoutExpired:
            return False, "Timeout."
        except Exception as e:
            return False, f"System error: {e}"

    def images_to_pdf(self, image_paths: List[str], pdf_path: str) -> Tuple[bool, str]:
        if not self.weasyprint_exe_path:
            return False, "weasyprint.exe not set."
        lines = ['<html><body style="margin:0;">']
        for img in image_paths:
            abs_path = os.path.abspath(img).replace('\\', '/')
            lines.append(f'<img src="file:///{abs_path}" style="width:100%; page-break-after:always;">')
        lines.append('</body></html>')
        html_content = '\n'.join(lines)
        temp_html = os.path.join(self.temp_dir, "temp_images.html")
        with open(temp_html, "w", encoding="utf-8") as f:
            f.write(html_content)
        success, msg = self.html_to_pdf(temp_html, pdf_path)
        try:
            os.remove(temp_html)
        except:
            pass
        return success, msg

    def cleanup_temp(self):
        try:
            shutil.rmtree(self.temp_dir, ignore_errors=True)
        except:
            pass

# ==================== UI Helpers ====================
class ProgressDialog:
    def __init__(self, parent, title: str, total: int):
        self.top = tk.Toplevel(parent)
        self.top.title(title)
        self.top.geometry("400x150")
        self.top.transient(parent)
        self.top.grab_set()
        self.top.update_idletasks()
        x = (parent.winfo_width() - 400)//2 + parent.winfo_x()
        y = (parent.winfo_height() - 150)//2 + parent.winfo_y()
        self.top.geometry(f"+{x}+{y}")
        ttk.Label(self.top, text="Processing...", font=('Segoe UI',11)).pack(pady=10)
        self.progress = ttk.Progressbar(self.top, orient="horizontal", length=350, mode="determinate")
        self.progress.pack(pady=10)
        self.progress["maximum"] = total
        self.status = ttk.Label(self.top, text="", wraplength=350)
        self.status.pack(pady=5)
        self.cancel = False
        ttk.Button(self.top, text="Cancel", command=self.on_cancel).pack(pady=5)

    def update(self, value: int, status: str = ""):
        self.progress["value"] = value
        if status:
            self.status.config(text=status)
        self.top.update()

    def on_cancel(self):
        self.cancel = True
        self.top.destroy()

class DarkTheme:
    @staticmethod
    def apply(root: tk.Tk) -> ttk.Style:
        style = ttk.Style(root)
        bg = "#1a1a1a"
        bg2 = "#2d2d2d"
        fg = "#ffffff"
        accent = "#58013B"
        root.configure(bg=bg)
        for theme in ("clam","alt","vista","xpnative","default"):
            try:
                style.theme_use(theme)
                break
            except:
                pass
        style.configure(".", background=bg, foreground=fg)
        style.configure("TFrame", background=bg)
        style.configure("TLabel", background=bg, foreground=fg)
        style.configure("Title.TLabel", background=bg, foreground=fg, font=("Segoe UI",11,"bold"))
        style.configure("TButton", background=bg2, foreground=fg, padding=8)
        style.map("TButton", background=[("active", accent)])
        style.configure("Primary.TButton", background=accent, foreground=fg, padding=9)
        style.map("Primary.TButton", background=[("active", "#7a1f5e")])
        style.configure("TEntry", fieldbackground=bg2, foreground=fg, insertcolor=fg, padding=6)
        style.configure("TCombobox", fieldbackground=bg2, foreground=fg, padding=6)
        style.configure("Sidebar.TFrame", background="#252526")
        style.configure("Card.TLabelframe", background=bg2, foreground=fg, bordercolor=accent)
        style.configure("Card.TLabelframe.Label", background=bg2, foreground=fg, font=("Segoe UI",10,"bold"))
        return style

# ==================== Main Application ====================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Production Analyzer — PDF (A4 Landscape Report, Image-based Dashboard)")
        self.geometry("1400x880")
        self.minsize(1200,720)

        self.style = DarkTheme.apply(self)
        configure_matplotlib_persian_font()
        MatplotlibViz.apply_modern_style()

        self.pdf_exporter = PDFExporter()
        self.weasyprint_exe_path = tk.StringVar(value="")
        self._find_weasyprint_exe()

        self.filepath = tk.StringVar()
        self.sheet_mode = tk.StringVar(value="Auto (27..39 except 33)")
        self.search_var = tk.StringVar()

        self.reader: Optional[ExcelReader] = None
        self.repo: Optional[DataRepository] = None
        self._internal_keys: List[str] = []
        self._display_keys: List[str] = []
        self.selected_stuff_internal: Optional[str] = None

        self.last_dashboard_html: Optional[str] = None
        self.last_report_html: Optional[str] = None

        self._build_ui()

    def _find_weasyprint_exe(self):
        
        current_dir = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(current_dir,"\weasyprint\dist\weasyprint.exe")
        paths = [
            r"C:\weasyprint\dist\weasyprint.exe",
            r"C:\Program Files\WeasyPrint\weasyprint.exe",
            r"C:\Program Files (x86)\WeasyPrint\weasyprint.exe",
            os.path.join(os.path.expanduser("~"), "weasyprint.exe"),
        ]
        for p in paths:
            if os.path.exists(p):
                self.weasyprint_exe_path.set(p)
                self.pdf_exporter.set_weasyprint_exe(p)
                break

    def _build_ui(self):
        main = ttk.Frame(self)
        main.pack(fill="both", expand=True)

        # Top bar
        top = tk.Frame(main, bg="#58013B", height=50)
        top.pack(fill="x")
        top.pack_propagate(False)
        tk.Label(top, text="Production Analyzer", bg="#58013B", fg="#ffffff",
                 font=("Segoe UI",14,"bold")).pack(side="left", padx=20, pady=10)
        tk.Label(top, text="PDF with charts & landscape reports", bg="#58013B", fg="#e0e0e0",
                 font=("Segoe UI",10)).pack(side="left", padx=10, pady=10)

        content = ttk.Frame(main)
        content.pack(fill="both", expand=True, padx=12, pady=12)

        # File selection
        file_frame = ttk.Frame(content)
        file_frame.pack(fill="x", pady=(0,12))
        ttk.Label(file_frame, text="Dataset", style="Title.TLabel").pack(side="left", padx=(0,10))
        ttk.Entry(file_frame, textvariable=self.filepath, width=80).pack(side="left", padx=(0,8))
        ttk.Button(file_frame, text="Browse", command=self.browse_file, style="Primary.TButton").pack(side="left", padx=2)
        ttk.Button(file_frame, text="Load", command=self.load_data, style="Primary.TButton").pack(side="left", padx=2)
        ttk.Label(file_frame, text="Sheets").pack(side="left", padx=(20,10))
        ttk.Combobox(file_frame, textvariable=self.sheet_mode,
                     values=["Auto (27..39 except 33)", "All sheets in file"], state="readonly", width=26).pack(side="left")

        # Paned window
        paned = ttk.Panedwindow(content, orient="horizontal")
        paned.pack(fill="both", expand=True)

        # Sidebar
        sidebar = ttk.Frame(paned, style="Sidebar.TFrame", padding=12)
        paned.add(sidebar, weight=1)
        ttk.Label(sidebar, text="Stuff Keys", style="Title.TLabel", background="#252526").pack(anchor="w")

        # Search
        search_frame = ttk.Frame(sidebar, style="Sidebar.TFrame")
        search_frame.pack(fill="x", pady=(10,8))
        ttk.Entry(search_frame, textvariable=self.search_var).pack(side="left", fill="x", expand=True)
        ttk.Button(search_frame, text="Filter", command=self.apply_filter).pack(side="left", padx=(8,0))
        ttk.Button(search_frame, text="Reset", command=self.reset_filter).pack(side="left", padx=(6,0))

        # Listbox
        self.listbox = tk.Listbox(sidebar, height=20, activestyle="none", bg="#2d2d2d", fg="#ffffff",
                                   selectbackground="#58013B", selectforeground="#ffffff")
        try:
            self.listbox.config(justify="right")
        except:
            pass
        self.listbox.pack(fill="both", expand=True)
        self.listbox.bind("<<ListboxSelect>>", self.on_select)

        # Action buttons
        actions = ttk.Labelframe(sidebar, text="Visualization", padding=10, style="Card.TLabelframe")
        actions.pack(fill="x", pady=(10,0))
        ttk.Button(actions, text="Matplotlib: Scatter", command=self.show_scatter,
                   style="Primary.TButton").pack(fill="x", pady=4)
        ttk.Button(actions, text="Plotly: Dashboard", command=self.show_plotly_dashboard,
                   style="Primary.TButton").pack(fill="x", pady=4)
        ttk.Button(actions, text="Web Report", command=self.show_web_report_refill,
                   style="Primary.TButton").pack(fill="x", pady=4)

        # PDF section
        pdf_frame = ttk.Labelframe(sidebar, text="PDF Export (weasyprint.exe)", padding=10, style="Card.TLabelframe")
        pdf_frame.pack(fill="x", pady=(10,0))
        exe_frame = ttk.Frame(pdf_frame)
        exe_frame.pack(fill="x", pady=5)
        ttk.Label(exe_frame, text="weasyprint.exe:").pack(side="left")
        ttk.Entry(exe_frame, textvariable=self.weasyprint_exe_path, width=20).pack(side="left", padx=5)
        ttk.Button(exe_frame, text="Browse", command=self.browse_weasyprint_exe).pack(side="left")

        ttk.Button(pdf_frame, text="📄 Current Dashboard → PDF (with images)",
                   command=self.download_dashboard_pdf, style="Primary.TButton").pack(fill="x", pady=4)
        ttk.Button(pdf_frame, text="📄 Current Report → PDF (A4 Landscape)",
                   command=self.download_report_pdf, style="Primary.TButton").pack(fill="x", pady=4)
        ttk.Separator(pdf_frame, orient='horizontal').pack(fill="x", pady=8)
        ttk.Button(pdf_frame, text="📑 Export ALL Stuff to PDF",
                   command=self.batch_export_all, style="Primary.TButton").pack(fill="x", pady=4)
        ttk.Button(pdf_frame, text="📚 Export Selected Stuff to PDF",
                   command=self.batch_export_selected, style="Primary.TButton").pack(fill="x", pady=4)

        status = "✓ weasyprint.exe ready" if self.pdf_exporter.weasyprint_exe_path else "⚠ weasyprint.exe not found"
        color = "#4caf50" if self.pdf_exporter.weasyprint_exe_path else "#f44336"
        ttk.Label(pdf_frame, text=status, foreground=color).pack(fill="x", pady=(5,0))

        # Main content area with notebook
        content_area = ttk.Frame(paned, padding=12)
        paned.add(content_area, weight=3)
        self.notebook = ttk.Notebook(content_area)
        self.notebook.pack(fill="both", expand=True)

        # Output tab
        out_tab = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(out_tab, text="Output")
        self.text = tk.Text(out_tab, wrap="word", bg="#2d2d2d", fg="#ffffff", insertbackground="#ffffff")
        self.text.tag_configure("rtl", justify="right")
        self.text.pack(fill="both", expand=True)

        # Plot tab
        plot_tab = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(plot_tab, text="Plot (Matplotlib)")
        self.figure, self.ax = plt.subplots(figsize=(10,6))
        self.figure.patch.set_facecolor('#2d2d2d')
        self.ax.set_facecolor('#2d2d2d')
        self.ax.tick_params(colors='#ffffff')
        self.ax.xaxis.label.set_color('#ffffff')
        self.ax.yaxis.label.set_color('#ffffff')
        self.ax.title.set_color('#ffffff')
        self.canvas = FigureCanvasTkAgg(self.figure, master=plot_tab)
        self.canvas.get_tk_widget().pack(fill="both", expand=True)
        NavigationToolbar2Tk(self.canvas, plot_tab).update()

        # Status bar
        self.status = tk.StringVar(value="Ready. Install kaleido for PDF charts: pip install kaleido")
        status_bar = tk.Label(main, textvariable=self.status, anchor="w", padx=12, pady=6,
                              bg="#58013B", fg="#ffffff")
        status_bar.pack(fill="x")

    def browse_file(self):
        path = filedialog.askopenfilename(initialdir=os.path.expanduser("~"),
                                          title="Select Excel file",
                                          filetypes=[("Excel files", "*.xlsx *.xls *.xlsm"), ("All files", "*.*")])
        if path:
            self.filepath.set(path)
            self.status.set(f"Selected: {path}")

    def browse_weasyprint_exe(self):
        path = filedialog.askopenfilename(title="Select weasyprint.exe",
                                          filetypes=[("Executable files", "*.exe"), ("All files", "*.*")])
        if path:
            self.weasyprint_exe_path.set(path)
            if self.pdf_exporter.set_weasyprint_exe(path):
                self.status.set(f"weasyprint.exe set to: {path}")
            else:
                messagebox.showerror("Invalid File", "Not a valid executable.")

    def _get_sheet_list(self, reader: ExcelReader) -> List[str]:
        if self.sheet_mode.get().startswith("All"):
            return reader.list_sheets()
        wanted = [str(i) for i in range(27,40) if i != 33]
        present = set(reader.list_sheets())
        return [s for s in wanted if s in present]

    def load_data(self):
        path = self.filepath.get().strip()
        if not path:
            messagebox.showwarning("No file", "Select an Excel file first.")
            return
        if not os.path.exists(path):
            messagebox.showerror("Not found", "File not found.")
            return
        self.reader = ExcelReader(path)
        self.repo = DataRepository(self.reader, DatasetParser())
        sheets = self._get_sheet_list(self.reader)
        self.status.set(f"Loading sheets: {', '.join(sheets)}")
        self.update_idletasks()
        self.repo.load(sheets)
        keys = sorted(self.repo.per_stuff.keys(), key=str)
        self._populate_listbox(keys)
        self.status.set(f"Loaded {len(self.repo.df_all):,} rows, {len(keys)} Stuff keys")

    def _populate_listbox(self, internal_keys: List[str]):
        self._internal_keys = list(internal_keys)
        self._display_keys = [fix_rtl_display(k) for k in self._internal_keys]
        self.listbox.delete(0, tk.END)
        for d in self._display_keys:
            self.listbox.insert(tk.END, d)

    def apply_filter(self):
        if not self.repo:
            return
        q = self.search_var.get().strip().lower()
        if not q:
            return
        keys = sorted([k for k in self.repo.per_stuff.keys() if q in str(k).lower()], key=str)
        self._populate_listbox(keys)
        self.status.set(f"Filter: {len(keys)} matches")

    def reset_filter(self):
        self.search_var.set("")
        if self.repo:
            self._populate_listbox(sorted(self.repo.per_stuff.keys(), key=str))
            self.status.set("Filter reset")

    def on_select(self, event):
        sel = event.widget.curselection()
        if not sel or not self.repo:
            return
        idx = int(sel[0])
        self.selected_stuff_internal = self._internal_keys[idx]
        self.status.set(f"Selected: {fix_rtl_display(self.selected_stuff_internal)}")

    def _get_selected_df(self) -> Optional[pd.DataFrame]:
        if not self.repo or not self.selected_stuff_internal:
            return None
        return self.repo.per_stuff.get(self.selected_stuff_internal)

    def show_scatter(self):
        df = self._get_selected_df()
        if df is None or df.empty:
            messagebox.showwarning("Select", "Select a Stuff key first.")
            return
        title = f"{fix_rtl_display('Plan vs Actual')} — {fix_rtl_display(self.selected_stuff_internal)}"
        MatplotlibViz.scatter_plan_actual(self.ax, df, title)
        self.canvas.draw()
        self.notebook.select(1)

    def show_plotly_dashboard(self):
        df = self._get_selected_df()
        if df is None or df.empty:
            messagebox.showwarning("Select", "Select a Stuff key first.")
            return
        path = PlotlyViz.interactive_dashboard(df, str(self.selected_stuff_internal))
        open_plotly_in_app(path, title=f"{fix_rtl_display(self.selected_stuff_internal)} — Dashboard")
        self.last_dashboard_html = path
        self.status.set("Dashboard created. Use PDF button to export with charts (requires kaleido).")

    def show_web_report_refill(self):
        df = self._get_selected_df()
        if df is None or df.empty:
            messagebox.showwarning("Select", "Select a Stuff key first.")
            return
        abefore = Analyzer(df)
        t0,p0,_,_ = abefore.t_test_plan_vs_actual()
        corr0 = abefore.correlation()
        reg0 = abefore.regression_plan_to_actual()

        aafter = Analyzer(df)
        filled_sr = aafter.refill_success_rate()
        filled_act = aafter.refill_actual()
        t1,p1,_,_ = aafter.t_test_plan_vs_actual()
        corr1 = aafter.correlation()
        reg1 = aafter.regression_plan_to_actual()

        self.repo.per_stuff[self.selected_stuff_internal] = aafter.df

        html = ReportViz.to_html_report(self.selected_stuff_internal,
                                        {"t":t0,"p":p0}, {"t":t1,"p":p1},
                                        filled_sr, filled_act, corr0, corr1, reg0, reg1)
        path = ReportViz.save_report(html, str(self.selected_stuff_internal))
        open_plotly_in_app(path, title=f"{fix_rtl_display(self.selected_stuff_internal)} — Report")
        self.last_report_html = path
        self.status.set("Report created. PDF will be A4 Landscape.")

    def download_dashboard_pdf(self):
        if not self.selected_stuff_internal:
            messagebox.showwarning("No Selection", "Select a Stuff item first.")
            return
        df = self._get_selected_df()
        if df is None or df.empty:
            messagebox.showwarning("No Data", "No data for selected item.")
            return
        if not self.pdf_exporter.weasyprint_exe_path:
            messagebox.showwarning("weasyprint.exe missing", "Set path to weasyprint.exe first.")
            return
        if not _HAS_KALEIDO:
            messagebox.showerror("Kaleido missing", "Install kaleido: pip install kaleido")
            return

        pdf_path = filedialog.asksaveasfilename(defaultextension=".pdf",
                                                filetypes=[("PDF files","*.pdf")],
                                                initialfile=f"{self.selected_stuff_internal}_dashboard.pdf")
        if not pdf_path:
            return

        self.status.set("Generating images...")
        self.update_idletasks()
        try:
            images = PlotlyViz.static_dashboard_images(df, self.selected_stuff_internal, self.pdf_exporter.temp_dir)
            self.status.set("Creating PDF...")
            self.update_idletasks()
            ok, msg = self.pdf_exporter.images_to_pdf(images, pdf_path)
            for img in images:
                try: os.remove(img)
                except: pass
            if ok:
                self.status.set(f"PDF saved: {os.path.basename(pdf_path)}")
                messagebox.showinfo("Success", "Dashboard PDF created.")
                if messagebox.askyesno("Open PDF", "Open now?"):
                    try: os.startfile(pdf_path)
                    except: pass
            else:
                self.status.set("PDF failed")
                messagebox.showerror("Error", f"Failed: {msg}")
        except Exception as e:
            self.status.set("Error")
            messagebox.showerror("Error", str(e))

    def download_report_pdf(self):
        if not self.last_report_html or not os.path.exists(self.last_report_html):
            messagebox.showwarning("No Report", "Generate a report first.")
            return
        if not self.pdf_exporter.weasyprint_exe_path:
            messagebox.showwarning("weasyprint.exe missing", "Set weasyprint.exe path.")
            return
        pdf_path = filedialog.asksaveasfilename(defaultextension=".pdf",
                                                filetypes=[("PDF files","*.pdf")],
                                                initialfile=f"{self.selected_stuff_internal}_report.pdf")
        if not pdf_path:
            return
        self.status.set("Converting report to PDF...")
        self.update_idletasks()
        ok, msg = self.pdf_exporter.html_to_pdf(self.last_report_html, pdf_path)
        if ok:
            self.status.set(f"PDF saved: {os.path.basename(pdf_path)}")
            messagebox.showinfo("Success", "Report PDF created.")
            if messagebox.askyesno("Open PDF", "Open now?"):
                try: os.startfile(pdf_path)
                except: pass
        else:
            self.status.set("PDF failed")
            messagebox.showerror("Error", f"Failed: {msg}")

    def batch_export_all(self):
        if not self.repo or not self.repo.per_stuff:
            messagebox.showwarning("No Data", "Load data first.")
            return
        if not self.pdf_exporter.weasyprint_exe_path:
            messagebox.showwarning("weasyprint.exe missing", "Set weasyprint.exe path.")
            return
        if not _HAS_KALEIDO:
            messagebox.showerror("Kaleido missing", "Install kaleido: pip install kaleido")
            return

        out_dir = filedialog.askdirectory(title="Select output folder", initialdir=os.path.expanduser("~/Desktop"))
        if not out_dir:
            return

        total = len(self.repo.per_stuff) * 2
        prog = ProgressDialog(self, "Batch Export", total)
        success, failed = [], []
        current = 0

        for name, df in self.repo.per_stuff.items():
            if prog.cancel:
                break
            # Report
            prog.update(current, f"Processing {name} report...")
            ab = Analyzer(df)
            aa = Analyzer(df)
            filled_sr = aa.refill_success_rate()
            filled_act = aa.refill_actual()
            t0,p0,_,_ = ab.t_test_plan_vs_actual()
            t1,p1,_,_ = aa.t_test_plan_vs_actual()
            corr0 = ab.correlation()
            corr1 = aa.correlation()
            reg0 = ab.regression_plan_to_actual()
            reg1 = aa.regression_plan_to_actual()
            html = ReportViz.to_html_report(name,
                                            {"t":t0,"p":p0}, {"t":t1,"p":p1},
                                            filled_sr, filled_act, corr0, corr1, reg0, reg1)
            html_path = os.path.join(self.pdf_exporter.temp_dir, f"{name}_report.html")
            with open(html_path,"w",encoding="utf-8") as f:
                f.write(html)
            pdf_report = os.path.join(out_dir, f"{name}_report.pdf")
            ok, msg = self.pdf_exporter.html_to_pdf(html_path, pdf_report)
            if ok:
                success.append(pdf_report)
            else:
                failed.append((pdf_report, msg))
            current += 1

            # Dashboard
            prog.update(current, f"Processing {name} dashboard...")
            try:
                images = PlotlyViz.static_dashboard_images(df, name, self.pdf_exporter.temp_dir)
                pdf_dash = os.path.join(out_dir, f"{name}_dashboard.pdf")
                ok, msg = self.pdf_exporter.images_to_pdf(images, pdf_dash)
                if ok:
                    success.append(pdf_dash)
                else:
                    failed.append((pdf_dash, msg))
                for img in images:
                    try: os.remove(img)
                    except: pass
            except Exception as e:
                failed.append((f"{name}_dashboard.pdf", str(e)))
            current += 1

        prog.top.destroy()
        self.pdf_exporter.cleanup_temp()

        msg = f"Successful: {len(success)} files\n"
        if failed:
            msg += f"Failed: {len(failed)} files\n"
            for f,err in failed[:3]:
                msg += f" • {os.path.basename(f)}: {err[:50]}...\n"
        msg += f"\nLocation: {out_dir}"
        if success:
            messagebox.showinfo("Batch Export", msg)
            if messagebox.askyesno("Open Folder", "Open output folder?"):
                try: os.startfile(out_dir)
                except: pass
        else:
            messagebox.showerror("Export Failed", "No PDFs created.")

    def batch_export_selected(self):
        if not self.repo or not self.repo.per_stuff:
            messagebox.showwarning("No Data", "Load data first.")
            return
        sel = self.listbox.curselection()
        if not sel:
            messagebox.showwarning("No Selection", "Select items from list.")
            return
        selected = [self._internal_keys[i] for i in sel]

        if not self.pdf_exporter.weasyprint_exe_path:
            messagebox.showwarning("weasyprint.exe missing", "Set weasyprint.exe path.")
            return
        if not _HAS_KALEIDO:
            messagebox.showerror("Kaleido missing", "Install kaleido: pip install kaleido")
            return

        out_dir = filedialog.askdirectory(title="Select output folder", initialdir=os.path.expanduser("~/Desktop"))
        if not out_dir:
            return

        total = len(selected) * 2
        prog = ProgressDialog(self, f"Exporting {len(selected)} items", total)
        success, failed = [], []
        current = 0

        for name in selected:
            if name not in self.repo.per_stuff:
                continue
            df = self.repo.per_stuff[name]
            if prog.cancel:
                break
            # Report
            prog.update(current, f"Processing {name} report...")
            ab = Analyzer(df)
            aa = Analyzer(df)
            filled_sr = aa.refill_success_rate()
            filled_act = aa.refill_actual()
            t0,p0,_,_ = ab.t_test_plan_vs_actual()
            t1,p1,_,_ = aa.t_test_plan_vs_actual()
            corr0 = ab.correlation()
            corr1 = aa.correlation()
            reg0 = ab.regression_plan_to_actual()
            reg1 = aa.regression_plan_to_actual()
            html = ReportViz.to_html_report(name,
                                            {"t":t0,"p":p0}, {"t":t1,"p":p1},
                                            filled_sr, filled_act, corr0, corr1, reg0, reg1)
            html_path = os.path.join(self.pdf_exporter.temp_dir, f"{name}_report.html")
            with open(html_path,"w",encoding="utf-8") as f:
                f.write(html)
            pdf_report = os.path.join(out_dir, f"{name}_report.pdf")
            ok, msg = self.pdf_exporter.html_to_pdf(html_path, pdf_report)
            if ok:
                success.append(pdf_report)
            else:
                failed.append((pdf_report, msg))
            current += 1

            # Dashboard
            prog.update(current, f"Processing {name} dashboard...")
            try:
                images = PlotlyViz.static_dashboard_images(df, name, self.pdf_exporter.temp_dir)
                pdf_dash = os.path.join(out_dir, f"{name}_dashboard.pdf")
                ok, msg = self.pdf_exporter.images_to_pdf(images, pdf_dash)
                if ok:
                    success.append(pdf_dash)
                else:
                    failed.append((pdf_dash, msg))
                for img in images:
                    try: os.remove(img)
                    except: pass
            except Exception as e:
                failed.append((f"{name}_dashboard.pdf", str(e)))
            current += 1

        prog.top.destroy()
        self.pdf_exporter.cleanup_temp()

        msg = f"Successful: {len(success)} files\n"
        if failed:
            msg += f"Failed: {len(failed)} files\n"
            for f,err in failed[:3]:
                msg += f" • {os.path.basename(f)}: {err[:50]}...\n"
        msg += f"\nLocation: {out_dir}"
        if success:
            messagebox.showinfo("Export Complete", msg)
            if messagebox.askyesno("Open Folder", "Open output folder?"):
                try: os.startfile(out_dir)
                except: pass
        else:
            messagebox.showerror("Export Failed", "No PDFs created.")

if __name__ == "__main__":
    app = App()
    app.mainloop()

# I warn you that this program only works with the datasets in PreProcessedDataset folder
# Have a nice day (;